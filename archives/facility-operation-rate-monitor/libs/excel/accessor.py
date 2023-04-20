import pathlib

import openpyxl as xl
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from utils.exceptions import InvalidAddressError


class ExcelAccessor():
    """Excelに対する読み書き操作をするクラス。

    NOTE:
    openpyxlのrow, column index は1からスタートする。

    Instance variables
    ----------
    wb: openpyxl.workbook.Workbook
        操作対象のWorkbooxオブジェクト。
    ws: openpyxl.worksheet.worksheet.Worksheet
        現在アクティブになっているWorksheetオブジェクト。
        NOTE: 初期化時は最初のシートをアクティブにする
    excel_path: str
        操作対象のExcelのパス。
        Excelの再読み込みに使う。
    """
    def __init__(self, excel_path: str | pathlib.Path, data_only=False) -> None:
        """インスタンスの初期化。

        Parameters
        ----------
        excel_path : str | pathlib.Path
            対象とするExcelファイルのパス。
        data_only : bool, optional
            Trueだと数式を評価した値を読み取る事ができる, by default False
        """
        try:
            self.wb: Workbook = xl.load_workbook(excel_path, data_only=data_only)
        except FileNotFoundError:
            self.wb: Workbook = xl.Workbook()
        self.ws: Worksheet = self.wb.worksheets[0]
        self.excel_path: str = str(pathlib.Path(excel_path).resolve())

    @property
    def sheet_size(self) -> int:
        """シート数を取得。

        Returns
        -------
        int
            シート数。
        """
        return len(self.wb.worksheets)

    @property
    def current_sheet_title(self) -> bytes | str | int:
        """現在アクティブになっているシートのタイトルを取得。

        Returns
        -------
        bytes | str | int
            現在アクティブになっているシートのタイトル。
        """
        return self.ws.title

    def reload(self, data_only=False):
        """Excelファイルを再読み込みする。

        Parameters
        ----------
        data_only : bool, optional
            Trueだと数式を評価した値を読み取る事ができる, by default False
        """
        self.wb: Workbook = xl.load_workbook(self.excel_path, data_only=data_only)

    def save_as(self, excel_path: str):
        """指定したパスでExcelファイルを保存する。
        NOTE: 拡張子が必要。

        Parameters
        ----------
        excel_path : str
            保存先のパス。
        needs_evaluate : bool, optional
            数式評価の要/否。, by default False
            Trueの場合 self.__evaluateメソッド を実行する。
        """
        self.wb.save(excel_path)

    def overwrite(self):
        """Excelファイルを上書き保存する。

        Parameters
        ----------
        needs_evaluate : bool, optional
            数式評価の要/否。, by default False
            Trueの場合 self.__evaluateメソッド を実行する。
        """
        self.wb.save(self.excel_path)

    def change_active_worksheet(self, sheet_title=None, sheet_index=None):
        """アクティブなシートを切り替える。
        優先度: sheet_title > sheet_index

        Parameters
        ----------
        sheet_title : str, optional
            シートのタイトル, by default None
        sheet_index : int, optional
            シートのインデックス, by default None
        """
        if sheet_title:
            self.ws = self.wb[sheet_title]
        # if sheet_index は sheet_indexが 0 の時にfalseになるため注意
        elif sheet_index is not None:
            self.ws = self.wb.worksheets[sheet_index]
        else:
            return

    def add_sheet(self, sheet_title: str, change_active_sheet: bool = False) -> Worksheet:
        """シートを追加する(最後に)

        Parameters
        ----------
        sheet_title : str
            シートのタイトル

        Returns
        -------
        Worksheet
            追加したシート
        """
        ws = self.wb.create_sheet(title=sheet_title)

        if change_active_sheet:
            self.ws = ws

        return ws

    def copy_worksheet(
        self,
        to_sheet_title: str,
        from_sheet_title=None,
        from_sheet_index=None,
        change_active_sheet=False,
    ):
        """ワークシートをコピーする。
        NOTE: シートタイトルが優先。

        Parameters
        ----------
        to_sheet_title : str
            コピー先のシートタイトル。
        from_sheet_title : str, optional
            コピー元のシートタイトル, by default None
        from_sheet_index : int, optional
            コピー元シートのインデックス, by default None
        change_active_sheet : bool, optional
            アクティブなシートを変更するかどうか, by default False
        """
        if from_sheet_title:
            ws = self.wb[from_sheet_title]  # type: ignore
        elif from_sheet_index is not None:
            ws = self.wb.worksheets[from_sheet_index]

        to_ws = self.wb.copy_worksheet(ws)
        to_ws.title = to_sheet_title

        if change_active_sheet:
            self.ws = to_ws

        return to_ws

    def read_cell_value(self, row=1, column=1, address=None) -> str | int | float | None:
        """セル(1つ)の値を取得する。
        NOTE: インデックスは 1~ のため注意。
        NOTE: 優先度: address > row, column

        Parameters
        ----------
        row : int, optional
            rowのインデックス, by default 1
        column : int, optional
            columnのインデックス, by default 1
        address : str, optional
            セルの番地。, by default None

        Returns
        -------
        Unknown | date | str | int | None
            セルの値。
        """
        if address:
            return self.ws[address].value
        return self.ws.cell(row=row, column=column).value

    def write_cell(
        self,
        value: int | float | str,
        row: int | tuple[int, int] = 1,
        column: int | tuple[int, int] = 1,
        address : str | None = None
    ):
        """セルに値を書き込む。
        NOTE: インデックスは 1~ のため注意。
        NOTE: 優先度: address > row, column

        Parameters
        ----------
        value : int | float | str
            セルに書き込む値。
        row : int, optional
            rowのインデックス, by default 1
        column : int, optional
            columnのインデックス, by default 1
        address : str, optional
            セルの番地。, by default None
        """
        # 文字列でアドレスが指定された場合
        if address:
            if len(address) == 2:
                self.ws[address].value = value
            # A1:B3のような場合
            elif len(address) == 5:
                self.ws.merge_cells(address)
                top_left_cell = self.ws[address[:2]]
                top_left_cell.value = value
            else:
                raise InvalidAddressError(f"無効なアドレスです {address}")
        # 数値でアドレスが指定された場合
        else:
            # 全て数値の場合 = 1つのセル
            if type(row) == type(column) == int:
                self.ws.cell(row=row, column=column, value=value)
            # 数値以外(タプル)を吹く場合 = 結合
            else:
                # rowの判定
                if type(row) == tuple:
                    merge_start_row, merge_end_row = row
                elif type(row) == int:
                    merge_start_row = merge_end_row = row
                else:
                    raise InvalidAddressError(f"無効なアドレスです {row, column}")
                # columnの判定
                if type(column) == tuple:
                    merge_start_column, merge_end_column = column
                elif type(column) == int:
                    merge_start_column = merge_end_column = column
                else:
                    raise InvalidAddressError(f"無効なアドレスです {row, column}")

                self.ws.merge_cells(
                    start_row=merge_start_row, end_row=merge_end_row,
                    start_column=merge_start_column, end_column=merge_end_column
                )
                self.ws.cell(row=merge_start_row, column=merge_start_column, value=value)

    def write_values_in_axial_direction(
        self,
        values: list[int | float | str],
        start_index: int,
        base_axis_index: int,
        axis=0
    ):
        """指定軸方向に値を書き込む。

        Parameters
        ----------
        values : list[int  |  float  |  str]
            書き込む値の配列。
        start_index : int
            開始インデックス(直交軸)。
        base_axis_index : int
            主軸のインデックス。
        axis : int, optional
            軸方向。0 => 横軸, 1 => 縦軸, by default 0
        """
        # 横方向: row固定, columnインクリメント
        if axis == 0:
            for i, value in enumerate(values):
                self.ws.cell(
                    row=base_axis_index,
                    column=start_index + i,
                    value=value
                )
        # 縦方向: rowインクリメント, column固定
        if axis == 1:
            for i, value in enumerate(values):
                self.ws.cell(
                    row=start_index + i,
                    column=base_axis_index,
                    value=value
                )

    def remove_sheet(self, sheet_index=None, sheet_title=None):
        """シートを削除する。
        優先度: sheet_title > sheet_index

        Parameters
        ----------
        sheet_index : int, optional
            削除したいシートのインデックス, by default None
        sheet_title : str, optional
            削除したいシートのタイトル, by default None
        """
        if sheet_title:
            ws = self.wb[sheet_title]
        elif sheet_index is not None:
            ws = self.wb.worksheets[sheet_index]
        self.wb.remove(ws)
