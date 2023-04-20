"""Excelファイルの読み書きを担うモジュール
"""
import os

import openpyxl as xl
from openpyxl.chart import Reference, ScatterChart, Series
from openpyxl.drawing.image import Image
from openpyxl.utils.cell import column_index_from_string, get_column_letter
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from backend.excel import types
from backend.excel.attachment import AttachedImage
from common import utils


class ExcelAccessor:
    """Excelに対する読み書き操作をするクラス。

    NOTE:
    row, column の index は 1 からスタートする。
    シートのインデックスも 1 からスタートする。

    NOTE:
    read_only, write_only には対応していない。
    """

    def __init__(
        self,
        src_filepath: str,
        first_active_sheet: types.SheetKey = 1,
    ) -> None:
        self._filepath: str = src_filepath
        self._load_workbook(first_active_sheet)  # _load_workbook は filepath を利用するのでセット後に実行する
        # print(f"Open Workbook: {self._filepath}")

    def __del__(self) -> None:
        self._wb.close()
        # print(f"Close Workbook: {self._filepath}")

    def _load_workbook(self, active_sheet: types.SheetKey = 1) -> None:
        """self.wb, self.active_worksheet を更新する。(self.filepath, self._is_xlsm()を利用)"""
        # 存在する場合は load , 存在しなければ新規作成する
        if os.path.isfile(self._filepath):
            self._wb: Workbook = xl.load_workbook(self._filepath)
        else:
            self._wb = xl.Workbook()
        self._active_worksheet: Worksheet = self._get_worksheet(active_sheet)

    def _get_worksheet(self, sheet_key: types.SheetKey) -> Worksheet:
        # インデックス指定の場合
        if isinstance(sheet_key, int):
            if sheet_key <= 0:
                raise KeyError("シートのインデックスは 1以上 の値で設定してください")
            try:
                ws = self._wb.worksheets[sheet_key - 1]
            except IndexError:
                raise KeyError(f"対象のシートが見つかりませんでした index: {sheet_key}")

        # 文字列指定の場合
        else:
            try:
                ws = self._wb[sheet_key]
            except KeyError:
                raise KeyError(f"対象のシートが見つかりませんでした title: {sheet_key}")

        # HACK: read_only \ write_only モードには対応しないため型無視する
        return ws  # type: ignore

    # Private
    def _validate_new_sheet_title(self, new_sheet_title: str) -> bool:
        if new_sheet_title in self._wb.sheetnames:
            return False
        return True

    # Static
    @staticmethod
    def validate_column_string(column: str) -> bool:
        try:
            column_index_from_string(column)
        except (ValueError, AttributeError):
            return False
        return True

    @staticmethod
    def _column_to_index_if_string(column: types.ColumnKey) -> int:
        if isinstance(column, str):
            return column_index_from_string(column)
        else:
            return column

    @staticmethod
    def _column_to_string_if_index(column: types.ColumnKey) -> str:
        if isinstance(column, int):
            return get_column_letter(column)
        else:
            return column

    # Property
    @property
    def number_of_sheets(self) -> int:
        return len(self._wb.worksheets)

    @property
    def current_sheet_title(self) -> str:
        # HACK: bytes | str | int 型となっているが無視する
        return self._active_worksheet.title  # type: ignore

    @property
    def sheet_titles(self) -> list[str]:
        return self._wb.sheetnames

    @property
    def min_row(self) -> int:
        return self._active_worksheet.min_row

    @property
    def max_row(self) -> int:
        return self._active_worksheet.max_row

    @property
    def min_column(self) -> int:
        return self._active_worksheet.min_column

    @property
    def max_column(self) -> int:
        return self._active_worksheet.max_column

    def save_as(self, filepath: str) -> None:
        """Save the workbook as filepath."""
        self._wb.save(filepath)
        self._filepath = filepath
        self._load_workbook(self.current_sheet_title)

    def overwrite(self) -> None:
        """Overwrite the workbook"""
        self._wb.save(self._filepath)

    def copy_worksheet(
        self,
        new_sheet_title: str,
        src_sheet_key: types.SheetKey,
        change_active_sheet=False,
    ) -> Worksheet:
        """Copy a worksheet and change the title. If needed, also change active sheet."""
        src_ws = self._get_worksheet(src_sheet_key)
        new_ws = self._wb.copy_worksheet(src_ws)
        new_ws.title = new_sheet_title
        # new_ws.title = INVALID_TITLE_REGEX.sub(new_sheet_title, "")

        if change_active_sheet:
            self._active_worksheet = new_ws

        # HACK: read_only \ write_only モードには対応しないため型無視する
        return new_ws  # type: ignore

    def change_active_worksheet(self, sheet_key: types.SheetKey) -> None:
        """Change the active worksheet."""
        self._active_worksheet = self._get_worksheet(sheet_key)

    def remove_worksheet(self, sheet_key: types.SheetKey) -> None:
        """Remove the worksheet from workbook."""
        self._wb.remove(self._get_worksheet(sheet_key))

    def read_cell_value_by_index(self, row: int, column: int) -> types.CellValue:
        """Read cell value at row and column (integer index)"""
        return self._active_worksheet.cell(row=row, column=column).value

    def read_cell_value_by_coordinate(self, coordinate: str) -> types.CellValue:
        """Read cell value at specific string coordinate"""
        return self._active_worksheet[coordinate].value

    def read_row_values(
        self, row_index: int, begin_column: types.ColumnKey, end_column: types.ColumnKey
    ) -> list[types.CellValue]:
        """Read cell values in row direction"""
        # HACK: openpyxlは書式のみが設定されている場合でもデータが存在すると認識してしまう
        # HACK: Noneのときはスキップすることで対処する
        begin_column_index = self._column_to_index_if_string(begin_column)
        end_column_index = self._column_to_index_if_string(end_column)
        values: list[types.CellValue] = []
        for column_index in range(begin_column_index, end_column_index + 1):  # 1-based index なので +1 する
            cell_value = self.read_cell_value_by_index(row=row_index, column=column_index)
            if cell_value is None:
                continue
            values.append(cell_value)
        return values

    def read_column_values(self, column: types.ColumnKey, begin_row: int, end_row: int) -> list[types.CellValue]:
        """Read cell values in column direction"""
        # HACK: openpyxlは書式のみが設定されている場合でもデータが存在すると認識してしまう
        # HACK: Noneのときはスキップすることで対処する
        column_index = self._column_to_index_if_string(column)
        values: list[types.CellValue] = []
        for row_index in range(begin_row, end_row + 1):  # 1-based index なので+1する
            cell_value = self.read_cell_value_by_index(row=row_index, column=column_index)
            if cell_value is None:
                continue
            values.append(cell_value)
        return values

    def write_cell_by_index(self, row: int, column: int, value: types.CellValue) -> None:
        """Write value to cell at row and column (integer index)"""
        self._active_worksheet.cell(row=row, column=column, value=value)

    def write_cell_by_coordinate(self, coordinate: str, value: types.CellValue) -> None:
        """Write value to cell at specific string coordinate"""
        self._active_worksheet[coordinate].value = value

    def write_row(
        self,
        row_index: int,
        begin_column: types.ColumnKey,
        values: list[types.CellValue],
    ) -> None:
        """Write values in row direction"""
        begin_column_index = self._column_to_index_if_string(begin_column)

        [
            self._active_worksheet.cell(row=row_index, column=begin_column_index + i, value=value)
            for i, value in enumerate(values)
        ]

    def write_column(
        self,
        begin_row_index: int,
        column: types.ColumnKey,
        values: list[types.CellValue],
    ) -> None:
        """Write values in column direction"""
        column_index = self._column_to_index_if_string(column)

        [
            self._active_worksheet.cell(row=begin_row_index + i, column=column_index, value=value)
            for i, value in enumerate(values)
        ]

    def add_image(self, attached_img: AttachedImage) -> None:
        """アクティブなワークシートに画像を貼り付ける。

        Parameters
        ----------
        attached_img : AttachedImage
            An instance of AttachedImage
        """
        image_to_excel = Image(attached_img.filename)
        image_to_excel.width = attached_img.width  # type: ignore
        image_to_excel.height = attached_img.height  # type: ignore
        self._active_worksheet.add_image(image_to_excel, attached_img.anchor)

    def add_scatter_chart(self, graph_src: types.Graph):
        """散布図をアクティブなシートに追加する。
        references: https://www.shibutan-bloomers.com/python_libraly_openpyxl-11/3439/

        Parameters
        ----------
        graph_src : Graph
            Graph(独自定義のdataclass)クラスのインスタンス。
        """
        # Init objects of openpyxl.
        x_value = Reference(self._active_worksheet, **graph_src.x_axis.members_as_dict)
        primary_chart = ScatterChart()
        secondary_chart = ScatterChart()

        # Set values to primary data.
        for i, primary_data in enumerate(graph_src.primary_data_collection):
            primary_y_value = Reference(self._active_worksheet, **primary_data.members_as_dict)
            series = Series(primary_y_value, x_value, title_from_data=True)
            primary_chart.append(series)
            s = primary_chart.series[i]
            s.marker.symbol = "none"
        # Set values to secondary data.
        secondary_y_value = Reference(self._active_worksheet, **graph_src.secondary_data.members_as_dict)
        series = Series(secondary_y_value, x_value, title_from_data=True)
        secondary_chart.append(series)
        s = secondary_chart.series[0]
        s.marker.symbol = "none"

        # Set Chart layout settings.
        primary_chart.title = graph_src.title
        primary_chart.width = graph_src.width  # type: ignore
        primary_chart.height = graph_src.height
        primary_chart.x_axis.title = graph_src.x_axis_title
        primary_chart.y_axis.title = graph_src.primary_y_axis_title
        if graph_src.legend is None:
            primary_chart.legend = None
        else:
            primary_chart.legend.position = graph_src.legend
        secondary_chart.y_axis.title = graph_src.secondary_y_axis_title
        secondary_chart.y_axis.axId = 200
        # Display y_axis of the secondary chart on the far right end on the x-axis
        secondary_chart.y_axis.crosses = "max"
        secondary_chart.x_axis.majorGridlines = None
        secondary_chart.y_axis.majorGridlines = None

        # Scaling
        primary_chart.x_axis.tickLblPos = "low"
        secondary_chart.x_axis.tickLblPos = "low"
        x_axis_interval = utils.decide_graph_scale_interval(graph_src.x_axis_max_value)
        primary_chart.x_axis.scaling.min = 0
        primary_chart.x_axis.scaling.max = graph_src.x_axis_max_value
        primary_chart.x_axis.majorUnit = x_axis_interval
        secondary_chart.x_axis.scaling.min = 0
        secondary_chart.x_axis.scaling.max = graph_src.x_axis_max_value
        secondary_chart.x_axis.majorUnit = x_axis_interval

        primary_chart += secondary_chart

        self._active_worksheet.add_chart(primary_chart, graph_src.address)
