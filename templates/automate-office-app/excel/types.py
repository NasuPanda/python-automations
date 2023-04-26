"""型を定義するモジュール
"""
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta
from decimal import Decimal
from typing import Literal, TypeAlias

from openpyxl.utils.cell import get_column_letter

# Excel
SheetKey: TypeAlias = int | str
ColumnKey: TypeAlias = int | str
CellValue: TypeAlias = (
    int | float | Decimal | str | bytes | datetime | date | time | timedelta | None
)
SheetReadingDirection = Literal["row", "column"]


@dataclass
class References:
    min_column: int
    max_column: int
    min_row: int
    max_row: int

    @property
    def members_as_dict(self):
        # openpyxlのメソッドに合わせた形にしておく
        return {
            "min_col": self.min_column,
            "max_col": self.max_column,
            "min_row": self.min_row,
            "max_row": self.max_row,
        }


@dataclass
class Graph:
    title: str
    x_axis_title: str
    x_axis_max_value: float
    primary_y_axis_title: str
    secondary_y_axis_max_vale: float
    secondary_y_axis_min_vale: float
    secondary_y_axis_interval: float
    secondary_y_axis_title: str
    primary_data_collection: list[References]
    secondary_data: References
    x_axis: References
    address: str
    width: int | float = 18  # default (15cm)
    height: int | float = 10  # default (7cm)
    legend: None | str = "b"


def __get_anchor_from_index(row: int, column: int) -> str:
    """Get anchor letter from row and column integer index"""
    return get_column_letter(column) + str(row)


@dataclass
class AttachedImage:
    """Attached image in Excel.

    NOTE: Adjust width and height based on default cell w/h.
    NOTE: 画像の配置や大きさの決め方は `/common/constants` 参照

    Example:
        img.width = 72 * 10
        img.height = 25 * 13

    References:
        - https://openpyxl.readthedocs.io/en/stable/api/openpyxl.worksheet.worksheet.html?highlight=add_image#openpyxl.worksheet.worksheet.Worksheet.add_image
        - https://openpyxl.readthedocs.io/en/stable/api/openpyxl.drawing.image.html?highlight=drawing#openpyxl.drawing.image.Image

    Arguments:
        filename: str
            画像のパス
        width: int
            画像の幅
        height: int
            画像の高さ
        anchor: str
            画像の位置(ex: A1, T4, ...)
    """

    filename: str
    width: int | float
    height: int | float
    anchor: str

    def __init__(
        self,
        filename: str,
        width: int | float,
        height: int | float,
        row_index: int,
        column_index: int,
        anchor_letter: None | str = None,
    ) -> None:
        self.filename = filename
        self.width = width
        self.height = height
        if anchor_letter:
            self.anchor = anchor_letter
        else:
            self.anchor = __get_anchor_from_index(row=row_index, column=column_index)
