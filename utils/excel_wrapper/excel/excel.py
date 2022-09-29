import os

import openpyxl as xl
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.utils.cell import column_index_from_string
import win32com.client

from .types import SheetKey, CellValue


class ExcelAccessor:
    """Excelに対する読み書き操作をするクラス。

    NOTE:
    row, column の index は 1 からスタートする。
    シートのインデックスは 0 からスタートする。

    NOTE:
    read_only, write_only には対応していない。
    """

    def __init__(self, excel_path: str, first_active_sheet: SheetKey = 0) -> None:
        if os.path.isfile(excel_path):
            self.wb: Workbook = xl.load_workbook(excel_path)
        else:
            self.wb = xl.Workbook()

        self.active_worksheet: Worksheet = self._get_worksheet(first_active_sheet)
        self.excel_path: str = excel_path

    def _get_worksheet(self, sheet_key: SheetKey) -> Worksheet:
        # インデックス指定の場合
        if isinstance(sheet_key, int):
            try:
                ws = self.wb.worksheets[sheet_key]
            except IndexError:
                raise KeyError(f"対象のシートが見つかりませんでした。 index: {sheet_key}")

        # 文字列指定の場合
        else:
            try:
                ws = self.wb[sheet_key]
            except KeyError:
                raise KeyError(f"対象のシートが見つかりませんでした。 title: {sheet_key}")

        # HACK: read_only \ write_only モードには対応しないため型無視する
        return ws  # type: ignore

    def _validate_new_sheet_title(self, new_sheet_title: str) -> None:
        if new_sheet_title in self.wb.sheetnames:
            raise ValueError(f"{new_sheet_title} is already in use.")

    @property
    def number_of_sheets(self) -> int:
        return len(self.wb.worksheets)

    @property
    def current_sheet_title(self) -> bytes | str | int:
        return self.active_worksheet.title

    def change_active_worksheet(self, sheet_key: SheetKey) -> None:
        self.active_worksheet = self._get_worksheet(sheet_key)

    def rename_active_worksheet(self, new_sheet_title: str) -> None:
        self._validate_new_sheet_title(new_sheet_title)
        self.active_worksheet.title = new_sheet_title

    def add_sheet(self, sheet_title: str) -> Worksheet:
        """シートを追加する(挿入位置は末尾)"""
        self._validate_new_sheet_title(sheet_title)
        # HACK : read_only \ write_only モードには対応しないため型無視する
        return self.wb.create_sheet(title=sheet_title)  # type: ignore

    def copy_worksheet(
        self,
        src_sheet_key: SheetKey,
        destination_sheet_title: str,
        needs_change_active_worksheet: bool = False,
    ) -> Worksheet:
        """シートを追加する(挿入位置は末尾)"""
        src_ws = self._get_worksheet(src_sheet_key)
        destination_ws = self.wb.copy_worksheet(src_ws)
        destination_ws.title = destination_sheet_title

        if needs_change_active_worksheet:
            self.active_worksheet = destination_ws

        return destination_ws

    def remove_sheet(self, sheet_key: SheetKey) -> None:
        ws = self._get_worksheet(sheet_key)
        self.wb.remove(ws)

    def save_as(self, excel_path: str) -> None:
        self.wb.save(excel_path)

    def overwrite(self) -> None:
        self.wb.save(self.excel_path)

    def read_cell_by_index(self, row: int, column: int) -> Cell | MergedCell:
        # HACK : read_only \ write_only モードには対応しないため型無視する
        return self.active_worksheet.cell(row=row, column=column)  # type: ignore

    def read_cell_value_by_index(self, row: int, column: int) -> CellValue:
        return self.read_cell_by_index(row, column).value

    def read_cell_by_coordinate(self, coordinate: str) -> Cell | MergedCell:
        return self.active_worksheet[coordinate]

    def read_cell_value_by_coordinate(self, coordinate: str) -> CellValue:
        return self.read_cell_by_coordinate(coordinate).value

    def write_cell_by_index(self, row: int, column: int, value: CellValue) -> None:
        self.active_worksheet.cell(row=row, column=column, value=value)

    def write_cell_by_coordinate(self, coordinate: str, value: CellValue) -> None:
        self.active_worksheet[coordinate](value=value)

    def write_row(self, row_index: int, begin_column: int | str, values: list[CellValue]) -> None:
        if isinstance(begin_column, str):
            begin_column_index = column_index_from_string(begin_column)
        else:
            begin_column_index = begin_column

        [
            self.active_worksheet.cell(row=row_index, column=begin_column_index + i, value=value)
            for i, value in enumerate(values)
        ]

    def write_column(self, begin_row_index: int, column: int | str, values: list[CellValue]) -> None:
        if isinstance(column, str):
            column_index = column_index_from_string(column)
        else:
            column_index = column

        [
            self.active_worksheet.cell(row=begin_row_index + i, column=column_index, value=value)
            for i, value in enumerate(values)
        ]
