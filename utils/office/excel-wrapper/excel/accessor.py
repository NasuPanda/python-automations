import os
import pathlib
import re
from shutil import register_unpack_format

import openpyxl as xl
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.utils.cell import column_index_from_string, get_column_letter
import pythoncom
import win32com.client

from excel import types


def relative_to_abs(path: str) -> str:
    """相対パスを絶対パスに変換する。(相対パスだと pywin32 でエラーが出るため)

    Args:
        path (str): 変換前のファイルパス。

    Returns:
        str: 絶対パスに変換したファイルパス。
    """
    return str(pathlib.Path(path).resolve())


class ExcelAccessor:
    """Excelに対する読み書き操作をするクラス。

    NOTE:
    row, column の index は 1 からスタートする。
    シートのインデックスも 1 からスタートする。

    NOTE:
    read_only, write_only には対応していない。
    """

    # ex: Sub FooBar()
    REGEX_MACRO_LINE = re.compile(r"Sub \S+\(\)")
    # `Sub ` or `()` にマッチ
    REGEX_MACRO_NAME = re.compile(r"Sub |\(\)")

    def __init__(
        self, src_filepath: str, dst_filepath: str | None = None, first_active_sheet: types.SheetKey = 1
    ) -> None:
        self.filepath: str = relative_to_abs(src_filepath)
        self._load_workbook(first_active_sheet)  # _load_workbook は filepath を利用するのでセット後に実行する

        # src と dst が異なる場合は save_as で dst の実体を作成しておく(pywin32 経由で操作するため)
        if dst_filepath:
            self.save_as(dst_filepath)

    def _load_workbook(self, active_sheet: types.SheetKey = 1) -> None:
        """self.wb, self.active_worksheet を更新する。(self.filepath, self._is_xlsm()を利用)"""
        # 存在する場合は load , 存在しなければ新規作成する
        if os.path.isfile(self.filepath):
            self.wb: Workbook = xl.load_workbook(self.filepath, keep_vba=self.is_xlsm)
        else:
            self.wb = xl.Workbook()
        self.active_worksheet: Worksheet = self._get_worksheet(active_sheet)

    def _get_worksheet(self, sheet_key: types.SheetKey) -> Worksheet:
        # インデックス指定の場合
        if isinstance(sheet_key, int):
            if sheet_key <= 0:
                raise KeyError(f"シートのインデックスは 1以上 の値で設定してください")
            try:
                ws = self.wb.worksheets[sheet_key - 1]
            except IndexError:
                raise KeyError(f"対象のシートが見つかりませんでした index: {sheet_key}")

        # 文字列指定の場合
        else:
            try:
                ws = self.wb[sheet_key]
            except KeyError:
                raise KeyError(f"対象のシートが見つかりませんでした title: {sheet_key}")

        # HACK: read_only \ write_only モードには対応しないため型無視する
        return ws  # type: ignore

    def _validate_new_sheet_title(self, new_sheet_title: str) -> bool:
        if new_sheet_title in self.wb.sheetnames:
            return False
        return True

    @classmethod
    def get_macro_names_from_code(cls, macro_code: str) -> list[str] | None:
        """マクロのコードからマクロ名の一覧を取得する。

        Args:
            macro_code (str): マクロのコード。

        Returns:
            list[str] | None: マクロ名の一覧。
        """
        macro_names: list[str] = []
        # ex: ["Sub Foo()", "Sub Bar()"]
        macro_name_lines: list[str] = cls.REGEX_MACRO_LINE.findall(macro_code)

        if not macro_name_lines:
            return

        # ex: ["Foo", "Bar"]
        [macro_names.append(cls.REGEX_MACRO_NAME.sub("", line)) for line in macro_name_lines]
        return macro_names

    @staticmethod
    def validate_column_string(column: str) -> bool:
        try:
            column_index_from_string(column)
        except (ValueError, AttributeError):
            return False
        return True

    @staticmethod
    def _column_to_index_if_string(column: types.ColumnKey) -> int:
        if isinstance(column, str):
            return column_index_from_string(column)
        else:
            return column

    @staticmethod
    def _column_to_string_if_index(column: types.ColumnKey) -> str:
        if isinstance(column, int):
            return get_column_letter(column)
        else:
            return column

    @property
    def number_of_sheets(self) -> int:
        return len(self.wb.worksheets)

    @property
    def current_sheet_title(self) -> str:
        # HACK: bytes | str | int 型となっているが str なので無視する
        return self.active_worksheet.title  # type: ignore

    @property
    def sheet_titles(self) -> list[str]:
        return self.wb.sheetnames

    @property
    def min_row(self) -> int:
        return self.active_worksheet.min_row

    @property
    def max_row(self) -> int:
        return self.active_worksheet.max_row

    @property
    def min_column(self) -> int:
        return self.active_worksheet.min_column

    @property
    def max_column(self) -> int:
        return self.active_worksheet.max_column

    @property
    def is_xlsm(self) -> bool:
        """xlsmファイルかどうか判定する"""
        return pathlib.Path(self.filepath).suffix == ".xlsm"

    def save_as(self, filepath: str) -> None:
        self.wb.save(filepath)
        self.filepath = filepath
        self._load_workbook(self.current_sheet_title)

    def overwrite(self) -> None:
        self.wb.save(self.filepath)

    def to_xlsm_if_xlsx(self) -> None:
        if not self.is_xlsm:
            # 変更を保存しておく
            self.overwrite()

            xlsm_filepath = str(pathlib.Path(self.filepath).with_suffix(".xlsm").resolve())

            xl = win32com.client.Dispatch("Excel.Application")
            wb = xl.Workbooks.Open(self.filepath)
            wb.SaveAs(xlsm_filepath, FileFormat=win32com.client.constants.xlOpenXMLWorkbookMacroEnabled)
            xl.Quit()

            self.filepath = xlsm_filepath
            self._load_workbook(self.current_sheet_title)

    def change_active_worksheet(self, sheet_key: types.SheetKey) -> None:
        self.active_worksheet = self._get_worksheet(sheet_key)

    def rename_active_worksheet(self, new_sheet_title: str) -> None:
        if not self._validate_new_sheet_title(new_sheet_title):
            raise ValueError(f"new_sheet_title: {new_sheet_title} は既に使われています")
        self.active_worksheet.title = new_sheet_title

    def add_sheet(self, new_sheet_title: str) -> Worksheet:
        """シートを追加する(挿入位置は末尾)"""
        if not self._validate_new_sheet_title(new_sheet_title):
            raise ValueError(f"new_sheet_title: {new_sheet_title} は既に使われています")
        # HACK : read_only \ write_only モードには対応しないため型無視する
        return self.wb.create_sheet(title=new_sheet_title)  # type: ignore

    def copy_worksheet(
        self,
        src_sheet_key: types.SheetKey,
        dst_sheet_title: str | None = None,
        needs_change_active_worksheet: bool = False,
    ) -> Worksheet:
        """シートを追加する(挿入位置は末尾)"""
        # 変更を保存、アクティブなシートを保持しておく
        self.overwrite()
        active_sheet_title_before_copy = self.current_sheet_title

        xl_app = win32com.client.Dispatch("Excel.Application")
        wb = xl_app.Workbooks.Open(self.filepath)
        wb.Worksheets(src_sheet_key).Copy(After=wb.Worksheets(self.number_of_sheets))

        if dst_sheet_title:
            if not self._validate_new_sheet_title(dst_sheet_title):
                raise ValueError(f"new_sheet_title: {dst_sheet_title} は既に使われています")
            xl_app.ActiveSheet.Name = dst_sheet_title

        wb.Save()
        wb.Close(True)
        xl_app.Quit()

        if needs_change_active_worksheet:
            # NOTE: load 前なので number_of_sheets の値は実際より 1 小さい
            self._load_workbook(active_sheet=self.number_of_sheets + 1)
        else:
            self._load_workbook(active_sheet=active_sheet_title_before_copy)

        return self._get_worksheet(sheet_key=self.number_of_sheets)

    def remove_sheet(self, sheet_key: types.SheetKey) -> None:
        ws = self._get_worksheet(sheet_key)
        self.wb.remove(ws)

    def read_cell_value_by_index(self, row: int, column: int) -> types.CellValue:
        return self.active_worksheet.cell(row=row, column=column).value

    def read_cell_value_by_coordinate(self, coordinate: str) -> types.CellValue:
        return self.active_worksheet[coordinate].value

    def read_row_values(
        self, row_index: int, begin_column: types.ColumnKey, end_column: types.ColumnKey
    ) -> list[types.CellValue]:
        begin_column_index = self._column_to_index_if_string(begin_column)
        end_column_index = self._column_to_index_if_string(end_column)
        return [
            self.read_cell_value_by_index(row=row_index, column=column_index)
            for column_index in range(begin_column_index, end_column_index + 1)  # 1-based index なので +1 する
        ]

    def read_column_values(self, column: types.ColumnKey, begin_row: int, end_row: int) -> list[types.CellValue]:
        column_index = self._column_to_index_if_string(column)
        return [
            self.read_cell_value_by_index(row=row_index, column=column_index)
            for row_index in range(begin_row, end_row + 1)  # 1-based index なので +1 する
        ]

    def read_current_sheet(
        self, sheet_reading_direction: types.SheetReadingDirection = "row"
    ) -> list[list[types.CellValue]]:
        if sheet_reading_direction == "row":
            return [
                self.read_row_values(row_index=row_index, begin_column=self.min_column, end_column=self.max_column)
                for row_index in range(self.min_row, self.max_row + 1)
            ]
        elif sheet_reading_direction == "column":
            return [
                self.read_column_values(column_index, self.min_row, self.max_row)
                for column_index in range(self.min_column, self.max_column + 1)
            ]
        else:
            raise ValueError(f"{sheet_reading_direction}は無効な引数です 次のいずれかにして下さい: {types.SheetReadingDirection}")

    def read_all_sheets(
        self, sheet_reading_direction: types.SheetReadingDirection = "row"
    ) -> dict[str, list[list[types.CellValue]]]:
        current_active_worksheet_title = self.current_sheet_title
        all_sheets_titles_and_values: dict[str, list[list[types.CellValue]]] = {}

        for sheet_title in self.sheet_titles:
            self.change_active_worksheet(sheet_title)
            all_sheets_titles_and_values[sheet_title] = self.read_current_sheet(sheet_reading_direction)

        self.change_active_worksheet(current_active_worksheet_title)
        return all_sheets_titles_and_values

    def write_cell_by_index(self, row: int, column: int, value: types.CellValue) -> None:
        self.active_worksheet.cell(row=row, column=column, value=value)

    def write_cell_by_coordinate(self, coordinate: str, value: types.CellValue) -> None:
        self.active_worksheet[coordinate].value = value

    def write_row(self, row_index: int, begin_column: types.ColumnKey, values: list[types.CellValue]) -> None:
        begin_column_index = self._column_to_index_if_string(begin_column)

        [
            self.active_worksheet.cell(row=row_index, column=begin_column_index + i, value=value)
            for i, value in enumerate(values)
        ]

    def write_column(self, begin_row_index: int, column: types.ColumnKey, values: list[types.CellValue]) -> None:
        column_index = self._column_to_index_if_string(column)

        [
            self.active_worksheet.cell(row=begin_row_index + i, column=column_index, value=value)
            for i, value in enumerate(values)
        ]

    def add_reference(
        self,
        row: int,
        column: types.ColumnKey,
        reference_row: int,
        reference_column: types.ColumnKey,
        another_sheet: types.SheetKey | None = None,
    ) -> None:
        # ex: "A1"
        reference_cell = self._column_to_string_if_index(reference_column) + str(reference_row)
        if another_sheet:
            # ex : =AnotherSheet!A1
            another_sheet_title = self._get_worksheet(another_sheet).title
            reference_cell_value = f"={another_sheet_title}!{reference_cell}"
        else:
            # ex: =A1
            reference_cell_value = f"={reference_cell}"

        self.write_cell_by_index(row=row, column=self._column_to_index_if_string(column), value=reference_cell_value)

    def add_macro(self, macro_code: str, is_visible: bool = True) -> None:
        """Excelファイルにマクロを追加する。
        NOTE: 定義先は ThisWorkbook とするため、`VBProject.VBComponents(1)`にアクセスする。

        Args:
            macro_code (str): 追加したいマクロのコード。
            is_visible (bool, optional): 操作中のExcelを表示するかどうか。デフォルトは `True` 。
        """
        xl = win32com.client.Dispatch("Excel.Application")
        xl.Visible = is_visible
        wb = xl.Workbooks.Open(self.filepath)

        macro_names = self.get_macro_names_from_code(macro_code)
        if macro_names:
            macro_name = macro_names[0]
        # 存在しなかった場合 None を返す return
        else:
            print(f"マクロから名前を取得出来ませんでした {macro_code}")
            return

        code_module = wb.VBProject.VBComponents(1).CodeModule
        existing_macro_code = code_module.Lines(1, code_module.CountOfLines)
        existing_macro_names = self.get_macro_names_from_code(existing_macro_code)

        # 同名のマクロが存在する場合 return
        if existing_macro_names:
            if macro_name in existing_macro_names:
                print(f"{macro_name} という名前のマクロは既に存在します")
                return

        try:
            wb.VBProject.VBComponents(1).CodeModule.AddFromString(macro_code)
            wb.Save()
        except pythoncom.com_error as e:
            print(e.excepinfo[2])
            print(
                """\
    以下の手順により解決する可能性があります。
        1. 「ファイル」タブを選択
        2. 「オプション」を選択
        3. 「セキュリティ センター」を選択
        4. 「セキュリティ センターの設定」を選択
        5. 「マクロの設定」 を選択
        6. 「VBA プロジェクト オブジェクト モデルへのアクセスを信頼する」のチェックボックスをオンにする
    """
            )
        finally:
            wb.Close(True)
            xl.Quit()

    def exec_macro(self, macro_name: str, is_visible: bool = True) -> None:
        """Excelファイルのマクロを実行する。

        Args:
            macro_name (str): マクロの名前。
            is_visible (bool, optional): 操作中のExcelを表示するかどうか。デフォルトは `True` 。
        """
        xl = win32com.client.Dispatch("Excel.Application")
        xl.Visible = is_visible
        wb = xl.Workbooks.Open(self.filepath)

        xl.Application.Run(macro_name)
        wb.Save()
        wb.Close()
        xl.Application.Quit()
