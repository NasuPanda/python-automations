import os
import re

import openpyxl as xl
from openpyxl.utils.cell import column_index_from_string, get_column_letter
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from src.common import types


class ExcelAccessor:
    """Excelに対する読み書き操作をするクラス。

    NOTE:
    row, column の index は 1 からスタートする。
    シートのインデックスも 1 からスタートする。

    NOTE:
    read_only, write_only には対応していない。
    """

    # ex: Sub FooBar()
    REGEX_MACRO_LINE = re.compile(r"Sub \S+\(\)")
    # `Sub ` or `()` にマッチ
    REGEX_MACRO_NAME = re.compile(r"Sub |\(\)")

    def __init__(self, filepath: str, first_active_sheet: types.SheetKey = 1) -> None:
        self.filepath: str = filepath
        self._load_workbook(first_active_sheet)  # _load_workbook は filepath を利用するのでセット後に実行する

    def _load_workbook(self, active_sheet: types.SheetKey = 1) -> None:
        """self.wb, self.active_worksheet を更新する。(self.filepath, self._is_xlsm()を利用)"""
        # 存在する場合は load , 存在しなければ新規作成する
        if os.path.isfile(self.filepath):
            self.wb: Workbook = xl.load_workbook(self.filepath)
        else:
            self.wb = xl.Workbook()
        self.active_worksheet: Worksheet = self._get_worksheet(active_sheet)

    def _get_worksheet(self, sheet_key: types.SheetKey) -> Worksheet:
        # インデックス指定の場合
        if isinstance(sheet_key, int):
            if sheet_key <= 0:
                raise KeyError(f"シートのインデックスは 1以上 の値で設定してください")
            try:
                ws = self.wb.worksheets[sheet_key - 1]
            except IndexError:
                raise KeyError(f"対象のシートが見つかりませんでした index: {sheet_key}")

        # 文字列指定の場合
        else:
            try:
                ws = self.wb[sheet_key]
            except KeyError:
                raise KeyError(f"対象のシートが見つかりませんでした title: {sheet_key}")

        # HACK: read_only \ write_only モードには対応しないため型無視する
        return ws  # type: ignore

    def _validate_new_sheet_title(self, new_sheet_title: str) -> bool:
        if new_sheet_title in self.wb.sheetnames:
            return False
        return True

    @staticmethod
    def validate_column_string(column: str) -> bool:
        try:
            column_index_from_string(column)
        except (ValueError, AttributeError):
            return False
        return True

    @staticmethod
    def _column_to_index_if_string(column: types.ColumnKey) -> int:
        if isinstance(column, str):
            return column_index_from_string(column)
        else:
            return column

    @staticmethod
    def _column_to_string_if_index(column: types.ColumnKey) -> str:
        if isinstance(column, int):
            return get_column_letter(column)
        else:
            return column

    @property
    def number_of_sheets(self) -> int:
        return len(self.wb.worksheets)

    @property
    def current_sheet_title(self) -> str:
        # HACK: bytes | str | int 型となっているが str なので無視する
        return self.active_worksheet.title  # type: ignore

    @property
    def sheet_titles(self) -> list[str]:
        return self.wb.sheetnames

    @property
    def min_row(self) -> int:
        return self.active_worksheet.min_row

    @property
    def max_row(self) -> int:
        return self.active_worksheet.max_row

    @property
    def min_column(self) -> int:
        return self.active_worksheet.min_column

    @property
    def max_column(self) -> int:
        return self.active_worksheet.max_column

    def save_as(self, filepath: str) -> None:
        self.wb.save(filepath)
        self.filepath = filepath
        self._load_workbook(self.current_sheet_title)

    def overwrite(self) -> None:
        self.wb.save(self.filepath)

    def change_active_worksheet(self, sheet_key: types.SheetKey) -> None:
        self.active_worksheet = self._get_worksheet(sheet_key)

    def rename_active_worksheet(self, new_sheet_title: str) -> None:
        if not self._validate_new_sheet_title(new_sheet_title):
            raise ValueError(f"new_sheet_title: {new_sheet_title} は既に使われています")
        self.active_worksheet.title = new_sheet_title

    def add_sheet(self, new_sheet_title: str) -> Worksheet:
        """シートを追加する(挿入位置は末尾)"""
        if not self._validate_new_sheet_title(new_sheet_title):
            raise ValueError(f"new_sheet_title: {new_sheet_title} は既に使われています")
        # HACK : read_only \ write_only モードには対応しないため型無視する
        return self.wb.create_sheet(title=new_sheet_title)  # type: ignore

    def remove_sheet(self, sheet_key: types.SheetKey) -> None:
        ws = self._get_worksheet(sheet_key)
        self.wb.remove(ws)

    def read_cell_value_by_index(self, row: int, column: int) -> types.CellValue:
        return self.active_worksheet.cell(row=row, column=column).value

    def read_cell_value_by_coordinate(self, coordinate: str) -> types.CellValue:
        return self.active_worksheet[coordinate].value

    def read_row_values(
        self, row_index: int, begin_column: types.ColumnKey, end_column: types.ColumnKey
    ) -> list[types.CellValue]:
        begin_column_index = self._column_to_index_if_string(begin_column)
        end_column_index = self._column_to_index_if_string(end_column)
        return [
            self.read_cell_value_by_index(row=row_index, column=column_index)
            for column_index in range(begin_column_index, end_column_index + 1)  # 1-based index なので +1 する
        ]

    def read_column_values(self, column: types.ColumnKey, begin_row: int, end_row: int) -> list[types.CellValue]:
        column_index = self._column_to_index_if_string(column)
        return [
            self.read_cell_value_by_index(row=row_index, column=column_index)
            for row_index in range(begin_row, end_row + 1)  # 1-based index なので +1 する
        ]

    def read_current_sheet(
        self, sheet_reading_direction: types.SheetReadingDirection = "row"
    ) -> list[list[types.CellValue]]:
        if sheet_reading_direction == "row":
            return [
                self.read_row_values(row_index=row_index, begin_column=self.min_column, end_column=self.max_column)
                for row_index in range(self.min_row, self.max_row + 1)
            ]
        elif sheet_reading_direction == "column":
            return [
                self.read_column_values(column_index, self.min_row, self.max_row)
                for column_index in range(self.min_column, self.max_column + 1)
            ]
        else:
            raise ValueError(f"{sheet_reading_direction}は無効な引数です 次のいずれかにして下さい: {types.SheetReadingDirection}")

    def read_all_sheets(
        self, sheet_reading_direction: types.SheetReadingDirection = "row"
    ) -> dict[str, list[list[types.CellValue]]]:
        current_active_worksheet_title = self.current_sheet_title
        all_sheets_titles_and_values: dict[str, list[list[types.CellValue]]] = {}

        for sheet_title in self.sheet_titles:
            self.change_active_worksheet(sheet_title)
            all_sheets_titles_and_values[sheet_title] = self.read_current_sheet(sheet_reading_direction)

        self.change_active_worksheet(current_active_worksheet_title)
        return all_sheets_titles_and_values

    def write_cell_by_index(self, row: int, column: int, value: types.CellValue) -> None:
        self.active_worksheet.cell(row=row, column=column, value=value)

    def write_cell_by_coordinate(self, coordinate: str, value: types.CellValue) -> None:
        self.active_worksheet[coordinate].value = value

    def write_row(self, row_index: int, begin_column: types.ColumnKey, values: list[types.CellValue]) -> None:
        begin_column_index = self._column_to_index_if_string(begin_column)

        [
            self.active_worksheet.cell(row=row_index, column=begin_column_index + i, value=value)
            for i, value in enumerate(values)
        ]

    def write_column(self, begin_row_index: int, column: types.ColumnKey, values: list[types.CellValue]) -> None:
        column_index = self._column_to_index_if_string(column)

        [
            self.active_worksheet.cell(row=begin_row_index + i, column=column_index, value=value)
            for i, value in enumerate(values)
        ]

    def add_reference(
        self,
        row: int,
        column: types.ColumnKey,
        reference_row: int,
        reference_column: types.ColumnKey,
        another_sheet: types.SheetKey | None = None,
    ) -> None:
        # ex: "A1"
        reference_cell = self._column_to_string_if_index(reference_column) + str(reference_row)
        if another_sheet:
            # ex : =AnotherSheet!A1
            another_sheet_title = self._get_worksheet(another_sheet).title
            reference_cell_value = f"={another_sheet_title}!{reference_cell}"
        else:
            # ex: =A1
            reference_cell_value = f"={reference_cell}"

        self.write_cell_by_index(row=row, column=self._column_to_index_if_string(column), value=reference_cell_value)
