import pathlib

import win32com.client
import openpyxl as xl
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

import config
from src.models.excel.graph import Graph
from src.utils.exceptions import ArgumentError
from src.utils import helper as Helper


class ExcelAccessor():
    """Excelに対する読み書き操作をするクラス。

    NOTE:
    openpyxlのrow, column index は 1からスタートする。

    NOTE:
    数式の評価が必要な場合, load時のdata_only=Trueのみでは不十分。
    Excelクライアントにより開いた時に初めて式が評価, 値が保存されるため。
    win32com.clientを用いて openpyxl.save → Excelクライアントで開いてから保存することで対処可能。

    Instance variables
    ----------
    wb: openpyxl.workbook.Workbook
        操作対象のWorkbooxオブジェクト。
    ws: openpyxl.worksheet.worksheet.Worksheet
        現在アクティブになっているWorksheetオブジェクト。
        NOTE: 初期化時は最初のシートをアクティブにする
    excel_path: str
        操作対象のExcelのパス。
        Excelの再読み込みに使う。
    """
    def __init__(self, excel_path: str | pathlib.Path, data_only=False) -> None:
        """インスタンスの初期化。

        Parameters
        ----------
        excel_path : str | pathlib.Path
            対象とするExcelファイルのパス。
        data_only : bool, optional
            Trueだと数式を評価した値を読み取る事ができる, by default False
        """
        self.wb: Workbook = xl.load_workbook(excel_path, data_only=data_only)
        self.ws: Worksheet = self.wb.worksheets[0]
        self.excel_path: str = Helper.resolve_relative_path(excel_path, needs_cast_str=True)

    @property
    def sheet_size(self) -> int:
        """シート数を取得。

        Returns
        -------
        int
            シート数。
        """
        return len(self.wb.worksheets)

    @property
    def current_sheet_title(self) -> bytes | str | int:
        """現在アクティブになっているシートのタイトルを取得。

        Returns
        -------
        bytes | str | int
            現在アクティブになっているシートのタイトル。
        """
        return self.ws.title

    def reload(self, data_only=False):
        """Excelファイルを再読み込みする。

        Parameters
        ----------
        data_only : bool, optional
            Trueだと数式を評価した値を読み取る事ができる, by default False
        """
        self.wb: Workbook = xl.load_workbook(self.excel_path, data_only=data_only)

    def save_as(self, excel_path: str, needs_evaluate=False):
        """指定したパスでExcelファイルを保存する。
        NOTE: 拡張子が必要。

        Parameters
        ----------
        excel_path : str
            保存先のパス。
        needs_evaluate : bool, optional
            数式評価の要/否。, by default False
            Trueの場合 self.__evaluateメソッド を実行する。
        """
        self.wb.save(excel_path)

        if needs_evaluate:
            self.__evaluate(excel_path)

    def overwrite(self, needs_evaluate=False):
        """Excelファイルを上書き保存する。

        Parameters
        ----------
        needs_evaluate : bool, optional
            数式評価の要/否。, by default False
            Trueの場合 self.__evaluateメソッド を実行する。
        """
        self.wb.save(self.excel_path)

        if needs_evaluate:
            self.__evaluate(self.excel_path)

    def __evaluate(self, excel_path: str):
        """数式を評価する。
        ExcelClientによりExcelファイルを開く→閉じることで数式を評価, データがファイルに保存される。

        Parameters
        ----------
        excel_path : str
            Excelのパス
        """
        app = win32com.client.Dispatch("Excel.Application")
        app.Visible = False
        app.DisplayAlerts = False
        # 相対パスを解釈してくれないのでresolve()しておく
        wb = app.Workbooks.Open(Helper.resolve_relative_path(excel_path, needs_cast_str=True))
        wb.Save()
        app.Quit()

    def change_active_worksheet(self, sheet_title=None, sheet_index=None):
        """アクティブなシートを切り替える。
        優先度: sheet_title > sheet_index

        Parameters
        ----------
        sheet_title : str, optional
            シートのタイトル, by default None
        sheet_index : int, optional
            シートのインデックス, by default None
        """
        if sheet_title:
            self.ws = self.wb[sheet_title]
        # if sheet_index は sheet_indexが 0 の時にfalseになるため注意
        elif sheet_index is not None:
            self.ws = self.wb.worksheets[sheet_index]
        else:
            raise ArgumentError('有効な引数が渡されていません。')

    def add_sheet(self, sheet_title: str) -> Worksheet:
        """シートを追加する(最後に)

        Parameters
        ----------
        sheet_title : str
            シートのタイトル

        Returns
        -------
        Worksheet
            追加したシート
        """
        ws = self.wb.create_sheet(title=sheet_title)
        return ws

    def copy_worksheet(
        self,
        to_sheet_title: str,
        from_sheet_title=None,
        from_sheet_index=None,
        change_active_sheet=False,
    ):
        """ワークシートをコピーする。
        NOTE: シートタイトルが優先。

        Parameters
        ----------
        to_sheet_title : str
            コピー先のシートタイトル。
        from_sheet_title : str, optional
            コピー元のシートタイトル, by default None
        from_sheet_index : int, optional
            コピー元シートのインデックス, by default None
        change_active_sheet : bool, optional
            アクティブなシートを変更するかどうか, by default False
        """
        if from_sheet_title:
            ws = self.wb[from_sheet_title]  # type: ignore
        elif from_sheet_index is not None:
            ws = self.wb.worksheets[from_sheet_index]
        else:
            raise ArgumentError("有効な引数が渡されていません。")

        to_ws = self.wb.copy_worksheet(ws)
        to_ws.title = to_sheet_title

        if change_active_sheet:
            self.ws = to_ws

        return to_ws

    def add_scatter(self, graph_src: Graph):
        """散布図をアクティブなシートに追加する。
        references: https://www.shibutan-bloomers.com/python_libraly_openpyxl-11/3439/

        Parameters
        ----------
        graph_src : Graph
            Graph(独自定義のdataclass)クラスのインスタンス。
        """
        # レイアウトの設定
        chart = ScatterChart()
        chart.width = graph_src.width
        chart.height = graph_src.height
        chart.title = graph_src.title
        chart.x_axis.title = graph_src.x_axis_title
        chart.y_axis.title = graph_src.y_axis_title
        if graph_src.legend is None:
            chart.legend = None
        else:
            chart.legend.position = graph_src.legend

        # データの参照を定義, Chartオブジェクトへ追加
        x_values = Reference(self.ws, **graph_src.x_references.members_as_dict())
        y_values = Reference(self.ws, **graph_src.y_references.members_as_dict())
        series = Series(x_values, y_values)
        chart.series.append(series)

        self.ws.add_chart(chart, graph_src.address)

    def read_cell_value(self, row=1, column=1, address=None) -> str | int | float | None:
        """セル(1つ)の値を取得する。
        NOTE: インデックスは 1~ のため注意。
        NOTE: 優先度: address > row, column

        Parameters
        ----------
        row : int, optional
            rowのインデックス, by default 1
        column : int, optional
            columnのインデックス, by default 1
        address : str, optional
            セルの番地。, by default None

        Returns
        -------
        Unknown | date | str | int | None
            セルの値。
        """
        if address:
            return self.ws[address].value
        return self.ws.cell(row=row, column=column).value

    def write_cell(self, value: int | float | str, row=1, column=1, address=None):
        """セルに値を書き込む。
        NOTE: インデックスは 1~ のため注意。
        NOTE: 優先度: address > row, column

        Parameters
        ----------
        value : int | float | str
            セルに書き込む値。
        row : int, optional
            rowのインデックス, by default 1
        column : int, optional
            columnのインデックス, by default 1
        address : str, optional
            セルの番地。, by default None
        """
        if address:
            self.ws[address].value = value
        else:
            self.ws.cell(row=row, column=column, value=value)

    def write_values_in_axial_direction(
        self,
        values: list[int | float | str],
        start_index: int,
        base_axis_index: int,
        axis=0
    ):
        """指定軸方向に値を書き込む。

        Parameters
        ----------
        values : list[int  |  float  |  str]
            書き込む値の配列。
        start_index : int
            開始インデックス(直交軸)。
        base_axis_index : int
            主軸のインデックス。
        axis : int, optional
            軸方向。0 => 縦軸, 1 => 横軸, by default 0
        """
        # 縦方向: rowインクリメント, column固定
        if axis == 0:
            for i, value in enumerate(values):
                self.ws.cell(
                    row=start_index + i,
                    column=base_axis_index,
                    value=value
                )
        # 横方向: row固定, columnインクリメント
        if axis == 1:
            for i, value in enumerate(values):
                self.ws.cell(
                    row=base_axis_index,
                    column=start_index + i,
                    value=value
                )

    def remove_sheet(self, sheet_index=None, sheet_title=None):
        """シートを削除する。
        優先度: sheet_title > sheet_index

        Parameters
        ----------
        sheet_index : int, optional
            削除したいシートのインデックス, by default None
        sheet_title : str, optional
            削除したいシートのタイトル, by default None
        """
        if sheet_title:
            ws = self.wb[sheet_title]
        elif sheet_index is not None:
            ws = self.wb.worksheets[sheet_index]
        else:
            raise ArgumentError("有効な引数が渡されていません。")
        self.wb.remove(ws)


class VIExcelAccessor(ExcelAccessor):
    """VIデータを扱うExcelAccessor。
    """
    # NOTE: 辞書として定義しているため ** を使って展開することでそのままメソッドに渡せる
    VALUE_CELL_ADDRESSES = config.EXCEL_VALUE_CELL_ADDRESSES
    DATA_START_ADDRESSES = config.EXCEL_DATA_START_ADDRESSES
    GRAPH = config.EXCEL_VI_GRAPH_CONFIG

    def __init__(self, excel_path: str | pathlib.Path, data_only=True) -> None:
        """インスタンスの初期化。

        Parameters
        ----------
        excel_path : str | pathlib.Path
            対象とするExcelファイルのパス。
        data_only : bool, optional
            Trueだと数式を評価した値を読み取る事ができる, by default False
        """
        super().__init__(excel_path, data_only)

    def add_vi_scatter(self):
        """散布図をアクティブなシートへ追加する。
        グラフレイアウト等は設定済。
        """
        self.add_scatter(self.GRAPH)
